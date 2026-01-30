import json
from itertools import count
from decimal import Decimal
from django.http import HttpResponse
from django.db.models import Sum, Avg, Max, Min, Q, Count
from rest_framework import viewsets, filters, status
from rest_framework.decorators import action, api_view
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .models import Region, Departement, Arrondissement, Production
from .serializers import (
    RegionSerializer, DepartementSerializer, 
    ArrondissementSerializer, ProductionSerializer,
    MapDataSerializer, AutocompleteSerializer
)


class RegionViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Region.objects.all().order_by('nom')
    serializer_class = RegionSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['nom', 'code']
    
    @action(detail=True, methods=['get'])
    def productions(self, request, pk=None):
        """Récupère toutes les productions d'une région"""
        region = self.get_object()
        productions = Production.objects.filter(region=region)
        serializer = ProductionSerializer(productions, many=True)
        return Response(serializer.data)


class DepartementViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Departement.objects.all().order_by('nom')
    serializer_class = DepartementSerializer
    filter_backends = [filters.SearchFilter, DjangoFilterBackend]
    search_fields = ['nom', 'code']
    filterset_fields = ['region']
    
    @action(detail=True, methods=['get'])
    def productions(self, request, pk=None):
        """Récupère toutes les productions d'un département"""
        departement = self.get_object()
        productions = Production.objects.filter(departement=departement)
        serializer = ProductionSerializer(productions, many=True)
        return Response(serializer.data)


class ArrondissementViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Arrondissement.objects.all().order_by('nom')
    serializer_class = ArrondissementSerializer
    filter_backends = [filters.SearchFilter, DjangoFilterBackend]
    search_fields = ['nom', 'code']
    filterset_fields = ['departement', 'departement__region']


class ProductionViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Production.objects.all().select_related(
        'region', 'departement', 'arrondissement'
    ).order_by('-annee', 'produit')
    serializer_class = ProductionSerializer
    filter_backends = [filters.SearchFilter, DjangoFilterBackend, filters.OrderingFilter]
    search_fields = ['produit', 'region__nom', 'departement__nom', 'arrondissement__nom']
    filterset_fields = ['secteur', 'annee', 'niveau_administratif', 'region', 'departement']
    ordering_fields = ['annee', 'quantite', 'produit']
    
    @action(detail=False, methods=['get'])
    def statistiques(self, request):
        """Retourne des statistiques filtrées sur les productions pour la synthèse"""
        queryset = self.filter_queryset(self.get_queryset())
        
        # Aggrégations de base
        total_productions = queryset.count()
        total_quantite = queryset.aggregate(total=Sum('quantite'))['total'] or 0
        
        # Par secteur pour trouver le secteur dominant
        par_secteur = queryset.values('secteur').annotate(
            count=Sum('quantite'),
            nombre=Count('id')
        ).order_by('-count')
        
        # Trouver la zone dominante (nom de la zone avec la plus grande production)
        # On regarde d'abord si un niveau administratif est filtré
        zone_dominante = "N/A"
        niveau = request.query_params.get('niveau_administratif')
        
        # On détermine la zone dominante en agrégeant par le niveau le plus précis disponible dans le filtrage
        if total_productions > 0:
            # On cherche la zone avec Max production dans le queryset filtré
            # Pour la synthèse simplifiée, on prend le premier résultat agrégé
            # (Plus complexe si on veut varier region/dept/arr, mais on va rester sur une approche robuste)
            top_record = queryset.order_by('-quantite').first()
            if top_record:
                zone_dominante = top_record.get_zone()

        return Response({
            'total_productions': total_productions,
            'total_quantite': float(total_quantite),
            'par_secteur': list(par_secteur),
            'zone_dominante': zone_dominante,
        })
    
    @action(detail=False, methods=['get'])
    def filtres(self, request):
        """Retourne les valeurs disponibles pour les filtres"""
        secteurs = Production.SECTEUR_CHOICES
        annees = Production.objects.values_list('annee', flat=True).distinct().order_by('-annee')
        produits = Production.objects.values_list('produit', flat=True).distinct().order_by('produit')
        
        return Response({
            'secteurs': [{'value': s[0], 'label': s[1]} for s in secteurs],
            'annees': list(annees),
            'produits': list(produits),
        })
    
    @action(detail=False, methods=['get'])
    def map_data(self, request):
        """
        Endpoint optimisé pour la carte interactive
        Retourne un GeoJSON avec les géométries et les données de production agrégées
        
        Paramètres:
        - secteur: agriculture, elevage, peche
        - produit: nom du produit
        - annee: année
        - niveau: region, departement, arrondissement
        """
        secteur = request.query_params.get('secteur')
        produit = request.query_params.get('produit')
        annee = request.query_params.get('annee')
        niveau = request.query_params.get('niveau', 'region')
        
        # Construire le filtre
        filters = {}
        if secteur:
            filters['secteur'] = secteur
        if produit:
            filters['produit'] = produit
        if annee:
            filters['annee'] = int(annee)
        filters['niveau_administratif'] = niveau
        
        # Récupérer les productions
        productions = Production.objects.filter(**filters)
        
        # Agréger par zone
        if niveau == 'region':
            aggregated = productions.values('region').annotate(
                total=Sum('quantite'),
                unite=Min('unite')
            )
            zones = Region.objects.filter(id__in=[p['region'] for p in aggregated if p['region']])
        elif niveau == 'departement':
            aggregated = productions.values('departement').annotate(
                total=Sum('quantite'),
                unite=Min('unite')
            )
            zones = Departement.objects.filter(id__in=[p['departement'] for p in aggregated if p['departement']])
        else:
            aggregated = productions.values('arrondissement').annotate(
                total=Sum('quantite'),
                unite=Min('unite')
            )
            zones = Arrondissement.objects.filter(id__in=[p['arrondissement'] for p in aggregated if p['arrondissement']])
        
        # Créer un dictionnaire des totaux
        totals_dict = {}
        unite_dict = {}
        for item in aggregated:
            zone_id = item.get('region') or item.get('departement') or item.get('arrondissement')
            if zone_id:
                totals_dict[zone_id] = float(item['total'])
                unite_dict[zone_id] = item['unite']
        
        # Construire le GeoJSON
        features = []
        for zone in zones:
            if not zone.geom_json:
                continue
            
            try:
                geometry = json.loads(zone.geom_json)
            except:
                continue
            
            total = totals_dict.get(zone.id, 0)
            unite = unite_dict.get(zone.id, '')
            
            feature = {
                'type': 'Feature',
                'id': zone.id,
                'properties': {
                    'id': zone.id,
                    'nom': zone.nom,
                    'code': zone.code,
                    'quantite': total,
                    'unite': unite,
                },
                'geometry': geometry
            }
            
            # Ajouter des infos hiérarchiques si nécessaire
            if niveau == 'departement':
                feature['properties']['region_nom'] = zone.region.nom
            elif niveau == 'arrondissement':
                feature['properties']['departement_nom'] = zone.departement.nom
                feature['properties']['region_nom'] = zone.departement.region.nom
            
            features.append(feature)
        
        # Calculer les métadonnées
        total_production = sum(totals_dict.values())
        zone_dominante = None
        max_production = 0
        
        if totals_dict:
            max_zone_id = max(totals_dict, key=totals_dict.get)
            max_production = totals_dict[max_zone_id]
            zone_dominante_obj = zones.filter(id=max_zone_id).first()
            if zone_dominante_obj:
                zone_dominante = zone_dominante_obj.nom
        
        metadata = {
            'secteur': secteur,
            'produit': produit,
            'annee': annee,
            'niveau': niveau,
            'total_production': total_production,
            'zone_dominante': zone_dominante,
            'production_max': max_production,
            'nombre_zones': len(features),
            'unite': unite_dict.get(list(unite_dict.keys())[0]) if unite_dict else '',
        }
        
        result = {
            'type': 'FeatureCollection',
            'features': features,
            'metadata': metadata
        }
        
        return Response(result)
    
    @action(detail=False, methods=['get'])
    def autocomplete(self, request):
        """
        Endpoint pour l'autocomplétion des lieux
        Paramètre: q (query string)
        """
        query = request.query_params.get('q', '').strip()
        
        if len(query) < 2:
            return Response([])
        
        results = []
        
        # Rechercher dans les régions
        regions = Region.objects.filter(nom__icontains=query)[:5]
        for region in regions:
            results.append({
                'id': region.id,
                'nom': region.nom,
                'type': 'region',
                'hierarchie': region.nom,
                'niveau_administratif': 'region'
            })
        
        # Rechercher dans les départements
        departements = Departement.objects.filter(nom__icontains=query).select_related('region')[:5]
        for dept in departements:
            results.append({
                'id': dept.id,
                'nom': dept.nom,
                'type': 'departement',
                'hierarchie': f"{dept.region.nom} > {dept.nom}",
                'niveau_administratif': 'departement'
            })
        
        # Rechercher dans les arrondissements
        arrondissements = Arrondissement.objects.filter(
            nom__icontains=query
        ).select_related('departement__region')[:5]
        for arr in arrondissements:
            results.append({
                'id': arr.id,
                'nom': arr.nom,
                'type': 'arrondissement',
                'hierarchie': f"{arr.departement.region.nom} > {arr.departement.nom} > {arr.nom}",
                'niveau_administratif': 'arrondissement'
            })
        
        # Limiter à 15 résultats
        return Response(results[:15])
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """
        Exporte les données de production en Excel
        Paramètres: secteur, produit, annee, niveau_administratif, region, departement
        """
        # Récupérer les filtres
        filters = {}
        if request.query_params.get('secteur'):
            filters['secteur'] = request.query_params.get('secteur')
        if request.query_params.get('produit'):
            filters['produit'] = request.query_params.get('produit')
        if request.query_params.get('annee'):
            filters['annee'] = int(request.query_params.get('annee'))
        if request.query_params.get('niveau_administratif'):
            filters['niveau_administratif'] = request.query_params.get('niveau_administratif')
        if request.query_params.get('region'):
            filters['region_id'] = int(request.query_params.get('region'))
        if request.query_params.get('departement'):
            filters['departement_id'] = int(request.query_params.get('departement'))
        
        # Récupérer les données
        productions = Production.objects.filter(**filters).select_related(
            'region', 'departement', 'arrondissement'
        ).order_by('-annee', 'secteur', 'produit')
        
        # Créer le workbook Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Productions"
        
        # Styles
        header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # En-têtes
        headers = [
            'Région', 'Département', 'Arrondissement', 'Secteur', 
            'Produit', 'Quantité', 'Unité', 'Année', 'Source'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Données
        for row_num, prod in enumerate(productions, 2):
            ws.cell(row=row_num, column=1, value=prod.region.nom if prod.region else '')
            ws.cell(row=row_num, column=2, value=prod.departement.nom if prod.departement else '')
            ws.cell(row=row_num, column=3, value=prod.arrondissement.nom if prod.arrondissement else '')
            ws.cell(row=row_num, column=4, value=prod.get_secteur_display())
            ws.cell(row=row_num, column=5, value=prod.produit)
            ws.cell(row=row_num, column=6, value=float(prod.quantite))
            ws.cell(row=row_num, column=7, value=prod.unite)
            ws.cell(row=row_num, column=8, value=prod.annee)
            ws.cell(row=row_num, column=9, value=prod.source_donnee)
        
        # Ajuster la largeur des colonnes
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Créer le nom du fichier dynamique
        secteur_str = request.query_params.get('secteur', 'tous')
        produit_str = request.query_params.get('produit', 'tous')
        annee_str = request.query_params.get('annee', 'toutes')
        
        filename = f"export_{secteur_str}_{produit_str}_{annee_str}_geoprod_cm.xlsx"
        # Remplacer les espaces par des underscores pour le nom du fichier
        filename = filename.replace(' ', '_').lower()

        # Créer la réponse HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        wb.save(response)
        return response